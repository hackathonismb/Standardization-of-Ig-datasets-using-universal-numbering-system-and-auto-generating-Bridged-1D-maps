from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import json
from colour import Color
from sys import argv
from copy import copy

colorcoded = int(argv[1])


if colorcoded:
	proteins = ["5ESV"]
	mapper = {"5ESV":{"light":"2","heavy":"1_contacts"}}
else:
	proteins = ["1RHH","2IG2","2ZJS","3DGG","4ZSO","5CMA","5ESV"]
	mapper = {	"1RHH":{"light":"1","heavy":"2"},
				"2IG2":{"light":"1","heavy":"2"},
				"2ZJS":{"light":"4","heavy":"3"},
				"3DGG":{"light":"1","heavy":"2"},
				"4ZSO":{"light":"1","heavy":"2"},
				"5CMA":{"light":"1","heavy":"2"},
				"5ESV":{"light":"2","heavy":"1"}}
col_range = {
	"heavy":['D','E','F','G','H','I','J','K','L','M','N','O','P'],
	"light":['X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']
	}
row_range = {
	"heavy":[str(i) for i in range(10,42)],
	"light":[str(i) for i in range(10,31)]
	}
properties = ["contacts"]
colors = [Color("black")]+list(Color("blue").range_to(Color("green"),15))
properties_mapper = {"contacts":{i:"FF"+colors[i].hex_l[1:].upper() for i in range(len(colors))}}


#for each protein
for protein in proteins:
	print(protein)
	if colorcoded:
		wb = load_workbook("igv_hl_template.xlsx")
	else:
		wb = load_workbook("igv_hl_template_default_colours.xlsx")
	ws = wb.active

#for both heavy and light chain
	#for chain in ["heavy","light"]:

	chain = "heavy"
	json_file = open("../data/input_files/"+protein+"_"+mapper[protein][chain]+".json")
	data = json.load(json_file)
		# create the empty excel sheet to write out
	col_out = 2
	wb_out = Workbook()
	ws_out = wb_out.active

#for each residue
	for res in data["residues"]: # for first
		if "kabat" not in res.keys():
			continue
		kabat = res["kabat"]
			
#search the template for the correct position
		for col in col_range[chain]:
			for row in row_range[chain]:
				cell = col+row
				if str(ws[cell].value) == kabat:

# fill the information

					for p in properties:
						if p in res.keys():
							pv = res[p]
							f = copy(ws[cell].font)
							f.color = properties_mapper[p][pv]
							ws_out.cell(row = 1, column = 1).value = 'kabad_num'
							ws_out.cell(row = 2, column = 1).value = protein+"_"+mapper[protein][chain]
							ws_out.cell(row = 3, column = 1).value = 'contacts'
							ws_out.cell(row = 1, column = col_out).value = res["kabat"]
							ws_out.cell(row = 2, column = col_out).value = res["res"]
							ws_out.cell(row = 2, column = col_out).font = f
							
							ws_out.cell(row = 2, column = col_out).fill = copy(ws[cell].fill)
							ws_out.cell(row = 3, column = col_out).value = res["contacts"]
						col_out += 1
	                                      	          

#save the 1D map
	if colorcoded:
		wb_out.save("../data/output2/"+protein+".xlsx")
	else:
		wb_out.save("../data/output/"+protein+".xlsx")
